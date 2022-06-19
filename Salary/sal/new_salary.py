import openpyxl
from openpyxl.styles import Font, Color
import pandas as pd
import xlwings as xw
import os
import smtplib

from email import encoders
from email.utils import formataddr
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
 
from_addr = formataddr(('오정훈', 'bst524@bigstonestudio.com'))

df = pd.DataFrame()
df= pd.read_excel('/Users/jonathanoh/Desktop/python/Salary/position.xlsx', index_col=0)


font = Font(name='Calibri',size=11,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='00FFFFFF')

salary = openpyxl.load_workbook('/Users/jonathanoh/Desktop/python/Salary/salary_datas.xlsx')
salary_sheet = salary.active


position = openpyxl.load_workbook('/Users/jonathanoh/Desktop/python/Salary/position.xlsx')
position_sheet = position.active


personal = openpyxl.load_workbook('/Users/jonathanoh/Desktop/python/Salary/personal.xlsx')
personal_sheet = personal.active

name_data =[]
# name_email =[]
for i in range(1, position_sheet.max_row):
    name_data.append(position_sheet.cell(row=i+1, column=1).value)
    # name_email.append(position_sheet.cell(row=i+1, column=5).value)

for names in name_data:
    for i in range(salary_sheet.max_row+1):
        for x in range(1, salary_sheet.max_column):    
            if salary_sheet.cell(row=i+1, column=1).value == names:
                personal_name=salary_sheet.cell(row=i+1, column=1).value
                personal_sheet.cell(row=1, column=x).value=salary_sheet.cell(row=i+1, column=x).value
                personal_sheet.cell(row=1, column=x).font = font
                personal_sheet.cell(row=2, column=x).value=salary_sheet.cell(row=i+2, column=x).value
                personal_sheet.cell(row=2, column=x).font = font
                personal_sheet.cell(row=1, column=salary_sheet.max_column+1).value=df.loc[personal_name][0]
                personal_sheet.cell(row=1, column=salary_sheet.max_column+1).font = font
                personal_sheet.cell(row=1, column=salary_sheet.max_column+2).value=df.loc[personal_name][1]
                personal_sheet.cell(row=1, column=salary_sheet.max_column+2).font = font
                personal_sheet.cell(row=1, column=salary_sheet.max_column+3).value=df.loc[personal_name][2]
                personal_sheet.cell(row=1, column=salary_sheet.max_column+3).font = font
            else :
                pass 
    name=salary_sheet.cell(row=i+1, column=1).value    
     
    personal.save('/Users/jonathanoh/Desktop/python/Salary/xltopdf/{0}.xlsx'.format(personal_name))
    wb=xw.Book('/Users/jonathanoh/Desktop/python/Salary/xltopdf/{0}.xlsx'.format(personal_name))
    wb.to_pdf('/Users/jonathanoh/Desktop/python/Salary/xltopdf/{0}.PDF'.format(personal_name))
    wb.close()
    
    
    to_addr =df.loc[personal_name][3]
    add_file = '/Users/jonathanoh/Desktop/python/Salary/xltopdf/{0}.PDF'.format(personal_name)
    
    session = None
    try:
        # SMTP 세션 생성
        session = smtplib.SMTP('smtp.gmail.com', 587)
        session.set_debuglevel(True)
        
        # SMTP 계정 인증 설정
        session.ehlo()
        session.starttls()
        session.login('j.h.empions@gmail.com', 'wyinguhzfbxulzvi')
    
        # 메일 콘텐츠 설정∏
        message = MIMEMultipart("mixed")
        
        # 메일 송/수신 옵션 설정
        message.set_charset('utf-8')
        message['From'] = from_addr
        message['To'] = to_addr
        message['Subject'] = str(personal_name) + '님 급여명세서' 
    
        # 메일 콘텐츠 - 내용
        body = '''
        <h4>안녕하세요.</h1>
        <h4>한달 동안 수고 많으셨습니다.</h1>
        <h4>급여 명세서 확인하십시오.</h1>
        <h4>오정훈 배상.</h1>
        '''
        bodyPart = MIMEText(body, 'html', 'utf-8')
        message.attach( bodyPart )
    
        # 메일 콘텐츠 - 첨부파일
        attachments = [
            # os.path.join( os.getcwd(), 'storage', '/Users/jonathanoh/Desktop/python/Salary/xltopdf/고지석_12월급여.pdf' )
            os.path.join(add_file)
        ]
    
        for attachment in attachments:
            attach_binary = MIMEBase("application", "octect-stream")
            try:
                binary = open(attachment, "rb").read() # read file to bytes
    
                attach_binary.set_payload( binary )
                encoders.encode_base64( attach_binary ) # Content-Transfer-Encoding: base64
                
                filename = os.path.basename( attachment )
                attach_binary.add_header("Content-Disposition", 'attachment', filename=('utf-8', '', filename))
                
                message.attach( attach_binary )
            except Exception as e:
                print( e )
    
        # 메일 발송
        session.sendmail(from_addr, to_addr, message.as_string())        
    
        print( 'Successfully sent the mail!!!' )
    except Exception as e:
        print( e )
    finally:
        if session is not None:
            session.quit()


