import tkinter as tk
from tkinter import CENTER, StringVar, ttk
from tkinter import filedialog
from typing_extensions import IntVar
import openpyxl
import pandas as pd
from openpyxl.styles import Font, Color
from regex import W
import xlwings as xw
import os
import smtplib

from email import encoders
from email.utils import formataddr
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

from yaml import safe_load
from_addr = formataddr(('오정훈', 'bst524@bigstonestudio.com'))

class FileSelection(tk.Frame):
    namedata= []
    peopledata =[]
    peopledatalist =[]
    salaryCost = []
    salaryCost2 = []
    salaryCostName = []
    errorname=[]
    font = Font(name='Calibri',size=11,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='00FFFFFF')

    
    def __init__(self, parent, *args, **kwargs):
        super().__init__(parent, *args, **kwargs)

        self.salaryMonth = tk.StringVar()
        self.salaryMonth.set("급여대장")
        self.salaryPeople = tk.StringVar()
        self.salaryPeople.set("급여 대상자")
        self.salaryPersonalSheet = tk.StringVar()
        self.salaryPersonalSheet.set("급여명세서 양식")
        self.start()

    def start(self):

        monthlySalary_button = ttk.Button(self, text="급여대장 선택", command = self.openSalarySheet )
        monthlySalary_label = ttk.Label(self, textvariable=self.salaryMonth)

        salaryPeople_button = ttk.Button(self, text="급여 대상자 명단  선택", command = self.openPeopleSheet)
        salaryPeople_label = ttk.Label(self, textvariable=self.salaryPeople)

        salaryPersonalSheet_button  = ttk.Button(self, text="급여 명세서 양식 선택", command = self.openPersonalSheet )
        salaryPersonalSheet_label = ttk.Label(self, textvariable=self.salaryPersonalSheet)

        emailbutton = ttk.Button(self, text = "급여 명세서 이메일 전송", command = self.makePersonalSheetEmail)
        personalSheet = ttk.Button(self, text = "급여 명세서 만들기", command =self.makePersonalSheet )
        selectAll = ttk.Button(self, text="select all", command=self.selectAll)
        deselectAll = ttk.Button(self, text= "deselect All", command=self.deselectAll)


        salaryPeople_button.grid(padx = 5,pady=5, row=0, column=0, sticky=(tk.W+tk.E))
        salaryPeople_label.grid(padx = 5,pady=5, row=0, column=1, sticky=(tk.W+tk.E))
        monthlySalary_button.grid(padx = 5,pady=5, row=1, column=0, sticky=(tk.W+tk.E))
        monthlySalary_label.grid(padx = 5, pady=5,  row=1, column=1, sticky=(tk.W+tk.E))
        salaryPersonalSheet_button.grid(padx = 5,pady=5, row=2, column=0, sticky=(tk.W+tk.E))
        salaryPersonalSheet_label.grid(padx = 5,pady=5, row=2, column=1, sticky=(tk.W+tk.E))
        emailbutton.grid(padx=5,pady=2, row=3, column=1, sticky=(tk.W+tk.E))
        personalSheet.grid(padx=5, pady=5, row=3, column =0, sticky=(tk.W+tk.E))
        selectAll.grid(padx=5, pady=5, row=3, column=2)
        deselectAll.grid(padx=5, pady=5, row=3, column=3)
    

    def openSalarySheet(self):
        if self.salaryCost:
            self.salaryCost=[]
            self.salaryCost2=[]
            self.salaryCostName=[]
            self.errornameLf.grid_forget()
        if self.namedata :
            self.salaryMonthFile = filedialog.askopenfilename(initialdir='/desktop/python/salary')
        # salaryMonthData = pd.read_excel(self.salaryMonthFile, engine = "openpyxl")
        # print(salaryMonthData)
            self.salaryMonth.set(self.salaryMonthFile.split("/")[-1])
            self.salaryMonthlyCostFile = openpyxl.load_workbook(self.salaryMonthFile)
            self.salaryMonthlyCostSheet = self.salaryMonthlyCostFile.active
            
            for name in self.namedata :
                aa=[]
                bb=[]
                for i in range(self.salaryMonthlyCostSheet.max_row+1):
                    if self.salaryMonthlyCostSheet.cell(row=i+1, column = 1).value == name :
                        for x in range(1,self.salaryMonthlyCostSheet.max_column):    
                            aa.append(self.salaryMonthlyCostSheet.cell(row=i+1, column=x).value)
                            if x == 1:
                                bb.append(self.salaryMonthlyCostSheet.cell(row=i+1, column=x).value)
                            else :
                                bb.append(self.salaryMonthlyCostSheet.cell(row=i+2, column=x).value)
                    elif self.salaryMonthlyCostSheet.cell(row=i+1, column = 2).value == name :
                        for x in range(2,self.salaryMonthlyCostSheet.max_column+1):    
                            aa.append(self.salaryMonthlyCostSheet.cell(row=i+1, column=x).value)
                            if x == 2:
                                bb.append(self.salaryMonthlyCostSheet.cell(row=i+1, column=x).value)
                            else :
                                bb.append(self.salaryMonthlyCostSheet.cell(row=i+2, column=x).value)
                        # elif self.salaryMonthlyCostSheet.cell(row=i+1, column = x).value == name :
                        #         # if self.salaryMonthlyCostSheet.cell(row=i+1, column =x).value :
                        #         # aa.append(self.salaryMonthlyCostSheet.cell(row=i+1, column=x).value)
                        #     aa.append(self.salaryMonthlyCostSheet.cell(row=i+1, column=x).value)
                        #     if x == 2:
                        #         bb.append(self.salaryMonthlyCostSheet.cell(row=i+1, column=x).value)
                        #     else :
                        #         bb.append(self.salaryMonthlyCostSheet.cell(row=i+2, column=x).value)
                if aa:
                    self.salaryCost.append(aa)
                    self.salaryCost2.append(bb) 
            for a in self.salaryCost:
                self.salaryCostName.append(a[0]) 
            
            self.errornameLf = tk.LabelFrame(self, padx=5, pady=5)
            self.errornameLf.grid(row=0,column=2)

            self.errorname=[]
            for a in self.namedata:
                if a in self.salaryCostName:
                    pass
                else : 
                    self.errorname.append(a)
            if self.errorname:
                tk.Label(self.errornameLf, text=self.errorname).grid(row=0, column=0)

        
    def openPeopleSheet(self):
        if self.peopledata:
            self.namedata=[]  
            self.peopledata=[]
            self.peopledatalist=[]
            self.chblabelframe.grid_forget()        
        self.salaryPeopleFile = filedialog.askopenfilename(initialdir='/desktop/python/salary')
        self.salaryPeople.set(self.salaryPeopleFile.split("/")[-1])
        self.peopleWb = openpyxl.load_workbook(self.salaryPeopleFile)
        self.peopleSheet = self.peopleWb.active
        for i in range(2, self.peopleSheet.max_row+1):
            if self.peopleSheet.cell(row=i, column=1).value != None:
                self.namedata.append(self.peopleSheet.cell(row=i, column=1).value)
                dd = (self.peopleSheet.cell(row=i, column=1).value,
                self.peopleSheet.cell(row=i, column=2).value, 
                self.peopleSheet.cell(row=i, column=3).value, 
                self.peopleSheet.cell(row=i, column=4).value,
                self.peopleSheet.cell(row=i, column=5).value, 
                self.peopleSheet.cell(row=i, column=6).value)
                self.peopledata.append(dd)
                bb=[]
                for d in range(1,7):
                    bb.append(self.peopleSheet.cell(row=i, column=d).value)
                self.peopledatalist.append(bb)
            else:
                pass
        print(self.namedata)
        self.cbText={}
        self.cbVariable={}
        self.cb={}
        self.cb_intvar=[]
        self.chblabelframe= tk.LabelFrame(self, padx=10, pady=5)
        self.chblabelframe.grid(row=4, column=0, columnspan=4, sticky=tk.W+tk.E)
        for idx, value in enumerate(self.peopledatalist):
            self.cb_intvar.append(tk.IntVar())

            if idx <= 6 :
                colval=0
                colval2 = idx
                
            elif idx<=13 :
                colval=1
                colval2 = idx-7
            elif idx<=20:
                colval=2
                colval2 = idx-14
            else :
                colval=3
                colval2 = idx-21
            tk.Checkbutton(self.chblabelframe, text=value[0], variable=self.cb_intvar[-1]).grid(row=colval2+4, column = colval, sticky='w')
            
            # self.cbText[idx] = StringVar()
            # self.cbText[idx].set(value[0])
            # self.cbVariable[idx] = IntVar()
            # self.cbVariable[idx].set(idx)
            # self.cb[idx]= ttk.Checkbutton(self, text=value[0], variable=self.cbVariable[idx], onvalue=1, offvalue=0, command=self.peopleSelect )
            # self.cb[idx].grid(row=idx+4,column=0, padx=5, pady=3)
    
    def deselectAll(self):
        for x in self.cb_intvar:
            x.set("0")
    
    def selectAll(self):
        for x in self.cb_intvar:
            x.set("1")
        #     print(t)
        # print(x.get() for x in self.cb_intvar)
        # pass

    def peopleSelect(self):
        self.selectedpeople=[]
        for ctr, intvar in enumerate(self.cb_intvar) :
            if intvar.get() :
                self.selectedpeople.append(self.namedata[ctr])
            else : 
                pass

    def peopleTable(self, peopledata):
        people_table= ttk.Treeview(self, column=[1,2,3,4,5,6])
        people_table_header = ["확인","이름","사번","부서","직위","이메일","지급일"]
        for idx, value in enumerate(people_table_header):
            people_table.heading('#{num}'.format(num=idx), text = value)
            people_table.column('#{num}'.format(num=idx), width=60, anchor=CENTER)
        people_table.column('#5', width=200)    
        people_table.grid(padx=10, pady=5, row=4, column=0, columnspan=4)
        for i in range(len(peopledata)):
            people_table.insert("", "end", text="", values=peopledata[i], iid=i)
        
    def openPersonalSheet(self):
        self.peopleSelect()
        self.personalSheetFile = filedialog.askopenfilename(initialdir='/desktop/python/salary')
        self.salaryPersonalSheet.set(self.personalSheetFile.split("/")[-1])
        self.personalSheetWb = openpyxl.load_workbook(self.personalSheetFile)
        self.personalSheetSh = self.personalSheetWb.active

    
    def makePersonalSheet(self):
        if self.errorname:
             pass
        else :

            self.peopleSelect()
            self.personalSheetWb = openpyxl.load_workbook(self.personalSheetFile)
            self.personalSheetSh = self.personalSheetWb.active

            for iidex,personaldata in enumerate (self.salaryCost):
                if personaldata[0] in self.selectedpeople:
                    for idx , personvalue in enumerate (personaldata) :
                        self.personalSheetSh.cell(row = 1, column = idx+1).value = personvalue
                        self.personalSheetSh.cell(row = 1, column = idx+1).font = self.font
                        self.personalSheetSh.cell(row = 1, column = len(personaldata)+2).value = self.peopledatalist[iidex][1]
                        self.personalSheetSh.cell(row = 1, column = len(personaldata)+2).font = self.font
                        self.personalSheetSh.cell(row = 1, column = len(personaldata)+3).value = self.peopledatalist[iidex][2]
                        self.personalSheetSh.cell(row = 1, column = len(personaldata)+3).font = self.font
                        self.personalSheetSh.cell(row = 1, column = len(personaldata)+4).value = self.peopledatalist[iidex][3]
                        self.personalSheetSh.cell(row = 1, column = len(personaldata)+4).font = self.font
                    emailaddr = self.peopledatalist[iidex][4]
                    # print(emailaddr)
                    for idx, personvalue in enumerate(self.salaryCost2[iidex]):
                        self.personalSheetSh.cell(row = 2, column = idx+1).value = personvalue
                        self.personalSheetSh.cell(row = 2, column = idx+1).font = self.font
                        
                    name = personaldata[0]
                    self.personalSheetWb.save('/Users/jonathanoh/Desktop/python/Salary/xltopdf/{0}.xlsx'.format(name))
                    wb=xw.Book('/Users/jonathanoh/Desktop/python/Salary/xltopdf/{0}.xlsx'.format(name))
                    wb.to_pdf('/Users/jonathanoh/Desktop/python/Salary/xltopdf/{0}.pdf'.format(name))
                    wb.close()
                    # self.mailSend(email=emailaddr, name=name)
                else:
                    pass

    def makePersonalSheetEmail(self):
        if self.errorname:
             pass
        else :

            self.peopleSelect()
            self.personalSheetWb = openpyxl.load_workbook(self.personalSheetFile)
            self.personalSheetSh = self.personalSheetWb.active

            for iidex,personaldata in enumerate (self.salaryCost):
                if personaldata[0] in self.selectedpeople:
                    for idx , personvalue in enumerate (personaldata) :
                        self.personalSheetSh.cell(row = 1, column = idx+1).value = personvalue
                        self.personalSheetSh.cell(row = 1, column = idx+1).font = self.font
                        self.personalSheetSh.cell(row = 1, column = len(personaldata)+2).value = self.peopledatalist[iidex][1]
                        self.personalSheetSh.cell(row = 1, column = len(personaldata)+2).font = self.font
                        self.personalSheetSh.cell(row = 1, column = len(personaldata)+3).value = self.peopledatalist[iidex][2]
                        self.personalSheetSh.cell(row = 1, column = len(personaldata)+3).font = self.font
                        self.personalSheetSh.cell(row = 1, column = len(personaldata)+4).value = self.peopledatalist[iidex][3]
                        self.personalSheetSh.cell(row = 1, column = len(personaldata)+4).font = self.font
                    emailaddr = self.peopledatalist[iidex][4]
                    # print(emailaddr)
                    for idx, personvalue in enumerate(self.salaryCost2[iidex]):
                        self.personalSheetSh.cell(row = 2, column = idx+1).value = personvalue
                        self.personalSheetSh.cell(row = 2, column = idx+1).font = self.font
                        
                    name = personaldata[0]
                    self.personalSheetWb.save('/Users/jonathanoh/Desktop/python/Salary/xltopdf/{0}.xlsx'.format(name))
                    wb=xw.Book('/Users/jonathanoh/Desktop/python/Salary/xltopdf/{0}.xlsx'.format(name))
                    wb.to_pdf('/Users/jonathanoh/Desktop/python/Salary/xltopdf/{0}.pdf'.format(name))
                    wb.close()
                    self.mailSend(email=emailaddr, name=name)
                else:
                    pass




    def mailSend(self, email, name ):
        
        to_addr =email
        add_file = '/Users/jonathanoh/Desktop/python/Salary/xltopdf/{0}.PDF'.format(name)
        
        session = None
        try:
            # SMTP 세션 생성
            session = smtplib.SMTP('smtp.gmail.com', 587)
            session.set_debuglevel(True)
            
            # SMTP 계정 인증 설정
            session.ehlo()
            session.starttls()
            session.login('j.h.empions@gmail.com', 'wyinguhzfbxulzvi')
        
            # 메일 콘텐츠 설정∏
            message = MIMEMultipart("mixed")
            
            # 메일 송/수신 옵션 설정
            message.set_charset('utf-8')
            message['From'] = from_addr
            message['To'] = to_addr
            message['Subject'] = str(name) + '님 급여명세서' 
        
            # 메일 콘텐츠 - 내용
            body = '''
            <h4>안녕하세요.</h1>
            <h4>한달 동안 수고 많으셨습니다.</h1>
            <h4>급여 명세서 확인하십시오.</h1>
            <h4>오정훈 배상.</h1>
            '''
            bodyPart = MIMEText(body, 'html', 'utf-8')
            message.attach( bodyPart )
        
            # 메일 콘텐츠 - 첨부파일
            attachments = [
                # os.path.join( os.getcwd(), 'storage', '/Users/jonathanoh/Desktop/python/Salary/xltopdf/고지석_12월급여.pdf' )
                os.path.join(add_file)
            ]
        
            for attachment in attachments:
                attach_binary = MIMEBase("application", "octect-stream")
                try:
                    binary = open(attachment, "rb").read() # read file to bytes
        
                    attach_binary.set_payload( binary )
                    encoders.encode_base64( attach_binary ) # Content-Transfer-Encoding: base64
                    
                    filename = os.path.basename( attachment )
                    attach_binary.add_header("Content-Disposition", 'attachment', filename=('utf-8', '', filename))
                    
                    message.attach( attach_binary )
                except Exception as e:
                    print( e )
        
            # 메일 발송
            session.sendmail(from_addr, to_addr, message.as_string())        
        
            print( 'Successfully sent the mail!!!' )
        except Exception as e:
            print( e )
        finally:
            if session is not None:
                session.quit()










    # def open():
    #     global my_image # 함수에서 이미지를 기억하도록 전역변수 선언 (안하면 사진이 안보임)
    #     root.filename = filedialog.askopenfilename(initialdir='', title='파일선택', filetypes=(
    # ('png files', '*.png'), ('jpg files', '*.jpg'), ('all files', '*.*')))
 
    #     Label(root, text=root.filename).pack() # 파일경로 view
    #     my_image = ImageTk.PhotoImage(Image.open(root.filename))
    #     Label(image=my_image).pack() #사진 view
 



# class FileMaking(tk.LabelFrame):
#     def __init__(self, parent, *args, **kwargs):
#         super().__init__(parent, *args, **kwargs)
#         treeview= ttk.Treeview(self, columns=["one", "two","three"], displaycolumns=["one","two","three"])
#         treeview.pack()



class MyApplication(tk.Tk):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.title("급여명세서 작성기")
        self.geometry("800x600")
        fk = FileSelection(self)
        fk.grid(sticky=(tk.E+tk.W+tk.N+tk.S))
        # FileMaking(self).grid(sticky=(tk.E+tk.W+tk.N+tk.S))
        


if __name__=='__main__':
    app = MyApplication()
    app.mainloop()
    