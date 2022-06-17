import openpyxl
from openpyxl.styles import Font, Color
import pandas as pd
import xlwings as xw
import os


df = pd.DataFrame()
df= pd.read_excel('/Users/jonathanoh/Desktop/python/Salary/position.xlsx', index_col=0)


font = Font(name='Calibri',size=11,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='00FFFFFF')

salary = openpyxl.load_workbook('/Users/jonathanoh/Desktop/python/Salary/salary_datas.xlsx')
salary_sheet = salary.active


position = openpyxl.load_workbook('/Users/jonathanoh/Desktop/python/Salary/position.xlsx')
position_sheet = position.active


personal = openpyxl.load_workbook('/Users/jonathanoh/Desktop/python/Salary/personal.xlsx')
personal_sheet = personal.active

name_data =[]
for i in range(1, position_sheet.max_row):
    name_data.append(position_sheet.cell(row=i+1, column=1).value)

for names in name_data:
    for i in range(salary_sheet.max_row+1):
        for x in range(1, salary_sheet.max_column):    
            if salary_sheet.cell(row=i+1, column=1).value == names:
                personal_name=salary_sheet.cell(row=i+1, column=1).value
                personal_sheet.cell(row=1, column=x).value=salary_sheet.cell(row=i+1, column=x).value
                personal_sheet.cell(row=1, column=x).font = font
                personal_sheet.cell(row=1, column=salary_sheet.max_column+1).value=df.loc[personal_name][0]
                personal_sheet.cell(row=1, column=salary_sheet.max_column+1).font = font
                personal_sheet.cell(row=1, column=salary_sheet.max_column+2).value=df.loc[personal_name][1]
                personal_sheet.cell(row=1, column=salary_sheet.max_column+2).font = font
                personal_sheet.cell(row=1, column=salary_sheet.max_column+3).value=df.loc[personal_name][2]
                personal_sheet.cell(row=1, column=salary_sheet.max_column+3).font = font
            else :
                pass 
    #name=salary_sheet.cell(row=i+1, column=1).value      
    personal.save('/Users/jonathanoh/Desktop/python/Salary/xltopdf/{0}.xlsx'.format(personal_name))
    wb=xw.Book('/Users/jonathanoh/Desktop/python/Salary/xltopdf/{0}.xlsx'.format(personal_name))
    wb.to_pdf('/Users/jonathanoh/Desktop/python/Salary/xltopdf/{0}.PDF'.format(personal_name))
    wb.close()
    
# import openpyxl


# wb= xw.Book('/Users/jonathanoh/Desktop/python/Salary/xltopdf/고지석.xlsx')
# ws= wb.sheets[0]

# current_work_dir = os.getcwd() 
# pdf_path = os.path.join(current_work_dir, "pdf파일명.pdf") 
# wb.to_pdf(pdf_path)
# # ws.api.ExportAsFixedFormat(0, pdf_path)


'''
for i in range(0,salary_sheet.max_row,3):
    for x in range(salary_sheet.max_column):
        if (salary_sheet.cell(row=i+1, column=1).value) is None or salary_sheet.cell(row=i+1, column=1).value=='':
            pass
        else :
            #print((sheet.cell(row=i+1, column=x+1).value))
            personal_sheet.cell(row=1, column=x+1).value = salary_sheet.cell(row=i+1, column=x+1).value
            personal_sheet.cell(row=1, column=x+1).font= font
    name=salary_sheet.cell(row=i+1, column=1).value

    personal.save('/Users/jonathanoh/Desktop/python/Salary/sal/{0}.xlsx'.format(name))

'''